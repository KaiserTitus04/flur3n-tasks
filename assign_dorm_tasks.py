#!/usr/bin/env python3
# assign_dorm_tasks.py  –  ✓ bleiben dauerhaft, alte Datei als Basis

import argparse, datetime as dt, re, random, sys
random.seed()
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

TEMPLATE_FILE = Path("DormTasks.xlsx")
PLANS_DIR     = Path("WeeklyPlans"); PLANS_DIR.mkdir(exist_ok=True)
DATE_RE       = re.compile(r"Tasks_(\d{4}-\d{2}-\d{2})\.xlsx")

# ───────── Excel‑Utilities ───────────────────────────────────────────────
def people_cols(ws) -> Dict[str, int]:
    mp, c = {}, 3
    while (v := ws.cell(row=4, column=c).value):
        mp[str(v).strip()] = c; c += 1
    return mp                      # {Name: Spalte}

def task_rows(ws) -> List[int]:
    rows, r = [], 5
    while ws.cell(row=r, column=2).value:
        rows.append(r); r += 1
    return rows

def latest_plan() -> Optional[Path]:
    dated = [(dt.date.fromisoformat(m.group(1)), p)
             for p in PLANS_DIR.glob("Tasks_*.xlsx")
             if (m := DATE_RE.match(p.name))]
    return max(dated, default=(None, None))[1] if dated else None

def next_meta_row(ws, start=17) -> int:
    r = start
    while ws.cell(row=r, column=2).value:
        r += 1
    return r

# ───────── Helper: alte X → ✓ ────────────────────────────────────────────
def convert_x_to_tick(ws, rows, pcols):
    for r in rows:
        for col in pcols.values():
            if str(ws.cell(row=r, column=col).value).upper() == "X":
                ws.cell(row=r, column=col, value="✓")

# ───────── Zuteilungs‑Funktionen ─────────────────────────────────────────
def start_pointer(last_ws, rows, pcols, order) -> int:
    if not last_ws:    # erster Lauf
        return 0
    first_row = rows[0]
    for i, name in enumerate(order):
        if str(last_ws.cell(row=first_row, column=pcols[name]).value).upper() == "X":
            return (i + 1) % len(order)
    return 0

def pick_candidate(cycle, excluded) -> Optional[str]:
    """liefert eine ZUFÄLLIGE Person aus cycle, die nicht in excluded ist."""
    pool = [p for p in cycle if p not in excluded]
    return random.choice(pool) if pool else None

# ───────── Haupt‑Routine ─────────────────────────────────────────────────
def build_plan(due: dt.date, absent: List[str], items: List[str]) -> Path:

    last_path = latest_plan()

    # Basis: letzter Plan, sonst Vorlage
    if last_path:
        wb = openpyxl.load_workbook(last_path)
    else:
        wb = openpyxl.load_workbook(TEMPLATE_FILE)

    ws       = wb.active
    pcols    = people_cols(ws)
    order    = list(pcols.keys())
    rows     = task_rows(ws)

    # (1) alte X in Basis‑Datei zu ✓ konvertieren
    convert_x_to_tick(ws, rows, pcols)

    # --- ZÄHLT, wie viele ✓ eine Person schon hat ----------------------------
    def task_counts(ws, rows, pcols) -> Dict[str, int]:
        counts = {name: 0 for name in pcols}
        for r in rows:
            for name, col in pcols.items():
                if str(ws.cell(row=r, column=col).value).strip() == "✓":
                    counts[name] += 1
        return counts
    
    # (1b) Summen der erledigten Tasks ermitteln
    done_counts = task_counts(ws, rows, pcols)


    # (2) Pointer auf Basis des _alten_ Plans bestimmen
    last_ws = openpyxl.load_workbook(last_path).active if last_path else None
    ptr     = start_pointer(last_ws, rows, pcols, order)

    cycle   = [p for p in order if p not in absent]
    weekly_assigned = set()
    rows_without_new_x = []

    for r in rows:
        # ► Abbruch, wenn mindestens eine Task leer blieb
        if rows_without_new_x:
            print("Keine neue Verteilung mehr möglich – Zeit für den nächsten Plan!", file=sys.stderr)
            sys.exit(1)
        done = {n for n, c in pcols.items()
                  if str(ws.cell(row=r, column=c).value).strip() == "✓"}
        excluded = done | weekly_assigned
        
        # --- Kandidat mit den wenigsten erledigten Aufgaben wählen --------
        pool = [p for p in cycle if p not in (done | weekly_assigned)]
        if pool:
            # kleinsten Zähler finden, bei Gleichstand per Zufall
            m = min(done_counts[p] for p in pool)
            cand = random.choice([p for p in pool if done_counts[p] == m])
            ws.cell(row=r, column=pcols[cand], value="X")
            weekly_assigned.add(cand)
            done_counts[cand] += 1          # Zähler sofort erhöhen
        else:
            rows_without_new_x.append(r)


    # Meta‑Daten
    mrow = next_meta_row(ws, 17)
    ws.cell(row=mrow, column=2, value=due.strftime("%Y-%m-%d"))
    for k, it in enumerate(items):
        ws.cell(row=mrow, column=9 + k, value=it.strip())

    out = PLANS_DIR / f"Tasks_{due.isoformat()}.xlsx"
    wb.save(out)

    # --- Kopie für GitHub Pages ---
    import shutil, pathlib
    latest = pathlib.Path("docs/WeeklyPlans/Tasks_latest.xlsx")
    latest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(out, latest)

    return out

# ───────── CLI ───────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--due", required=True, help="YYYY-MM-DD")
    ap.add_argument("--absent", default="", help="Komma‑Liste Abwesende")
    ap.add_argument("--buy",    default="", help="Komma‑Liste Einkäufe")
    a = ap.parse_args()

    due  = dt.date.fromisoformat(a.due)
    away = [s.strip() for s in a.absent.split(",") if s.strip()]
    buy  = [s.strip() for s in a.buy.split(",")    if s.strip()]

    path = build_plan(due, away, buy)
    print("✔ Plan erzeugt:", path)

if __name__ == "__main__":
    main()
